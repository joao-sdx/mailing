from __future__ import annotations

import sys
import time
import urllib.request
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).parent
EXCEL_FILE = SCRIPT_DIR / "credit-bail.xlsx"
API_URL = "https://recherche-entreprises.api.gouv.fr/search?q={siren}"
MAX_WORKERS = 10
RETRY = 3


def fetch(siren: str) -> dict | None:
    url = API_URL.format(siren=siren)
    for attempt in range(RETRY):
        try:
            with urllib.request.urlopen(url, timeout=10) as resp:
                data = json.loads(resp.read())
                results = data.get("results", [])
                return results[0] if results else None
        except Exception:
            if attempt == RETRY - 1:
                return None
            time.sleep(0.5)


def latest_finances(finances: dict | None) -> tuple:
    if not finances:
        return None, None
    latest = max(finances.keys())
    year = finances[latest]
    return year.get("ca"), year.get("resultat_net")


def enrich_all(sirens: list[str]) -> dict[str, dict]:
    results = {}
    total = len(sirens)
    done = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(fetch, s): s for s in sirens}
        for future in as_completed(futures):
            siren = futures[future]
            results[siren] = future.result()
            done += 1
            if done % 100 == 0 or done == total:
                print(f"  {done}/{total} SIRENs traités...", file=sys.stderr)

    return results


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_siren = wb["SIREN"]

    # Find siren column index
    headers = [ws_siren.cell(1, c).value for c in range(1, ws_siren.max_column + 1)]
    siren_col = headers.index("siren") + 1

    sirens = [
        str(ws_siren.cell(row, siren_col).value).strip()
        for row in range(2, ws_siren.max_row + 1)
    ]

    print(f"Enrichissement de {len(sirens)} SIRENs...")
    data = enrich_all(sirens)

    # Add new columns to SIREN sheet
    next_col = ws_siren.max_column + 1
    new_cols = ["site_web", "adresse_siege", "ca", "resultat_net"]
    for i, col_name in enumerate(new_cols):
        ws_siren.cell(1, next_col + i, col_name)

    col_site = next_col
    col_adresse = next_col + 1
    col_ca = next_col + 2
    col_resultat = next_col + 3

    # Contacts sheet
    if "Contacts" in wb.sheetnames:
        del wb["Contacts"]
    ws_contacts = wb.create_sheet("Contacts")
    ws_contacts.append(["siren", "nom", "prenom", "titre"])

    contact_count = 0
    for row in range(2, ws_siren.max_row + 1):
        siren = str(ws_siren.cell(row, siren_col).value).strip()
        result = data.get(siren)

        if result:
            adresse = result.get("siege", {}).get("adresse")
            ca, resultat_net = latest_finances(result.get("finances"))
            dirigeants = result.get("dirigeants") or []
        else:
            adresse = ca = resultat_net = None
            dirigeants = []

        ws_siren.cell(row, col_site, None)
        ws_siren.cell(row, col_adresse, adresse)
        ws_siren.cell(row, col_ca, ca)
        ws_siren.cell(row, col_resultat, resultat_net)

        for d in dirigeants:
            if d.get("type_dirigeant") == "personne physique":
                nom = d.get("nom", "")
                prenom = d.get("prenoms", "")
            else:
                nom = d.get("denomination", "")
                prenom = ""
            titre = d.get("qualite", "")
            ws_contacts.append([siren, nom, prenom, titre])
            contact_count += 1

    wb.save(EXCEL_FILE)
    print(f"Terminé : {len(sirens)} SIRENs enrichis, {contact_count} contacts → {EXCEL_FILE.name}")


if __name__ == "__main__":
    main()
